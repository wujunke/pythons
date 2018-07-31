

def makeExportOrgExcel(self, orgidlist):
    # 获取数据
    import openpyxl
    from openpyxl import styles
    org_qs = self.queryset.filter(id__in=orgidlist)
    path = ''

    class downloadAllDataroomFile(threading.Thread):
        def __init__(self, org_qs, path):
            self.org_qs = org_qs
            self.path = path
            threading.Thread.__init__(self)

        def run(self):
            wb = openpyxl.Workbook()
            ws_org = wb.active
            ws_org.title = u'机构列表'
            # ws_org = wb.create_sheet(title=u'机构列表')
            ws_org['A1']= u'机构全称'
            ws_org['B1'] = u'描述'
            ws_org['C1'] = u'合伙人/投委会成员'
            ws_org['D1'] = u'标签'
            ws_org['E1'] = u'投资事件'
            ws_org.column_dimensions['B'].width = 50
            ws_org.column_dimensions['E'].width = 120
            ws_org_hang = 2
            sheet_index = 0
            for org in org_qs:
                tags = org.tags.all()
                tagnamelist = []
                for tag in tags:
                    tagnamelist.append(tag.nameC)
                tagstr = '、'.join(tagnamelist)
                investevents = org.org_orgInvestEvent.all().filter(is_deleted=False)
                event_list = []
                com_list = []
                for event in investevents:
                    com_list.append(event.com_id)
                    event_list.append('项目名称：%s, 行业：%s , 投资日期：%s , 投资轮次：%s, 投资金额：%s' %
                                      (event.comshortname, event.industrytype, str(event.investDate)[:10],
                                       event.investType if event.investType else '暂无',
                                       event.investSize if event.investSize else '暂无'))
                eventstr = '\n\r'.join(event_list)
                ws_org['A%s' % ws_org_hang] = str(org.orgfullname)
                ws_org['B%s' % ws_org_hang] = str(org.description) if org.description else '暂无'
                ws_org['C%s' % ws_org_hang] = str(org.partnerOrInvestmentCommiterMember) if org.partnerOrInvestmentCommiterMember else '暂无'
                ws_org['D%s' % ws_org_hang] = tagstr if len(tagstr) > 0 else '暂无'
                ws_org['E%s' % ws_org_hang] = eventstr if len(eventstr) > 0 else '暂无'

                com_list = ProjectData.objects.filter(com_id__in=com_list)
                if len(com_list) > 0:
                    sheet_index = sheet_index + 1
                    com_sheet = wb.create_sheet(org.orgfullname, index=sheet_index)
                    com_sheet['A1'] = u'全称'
                    com_sheet['B1'] = u'简介'
                    com_sheet['C1'] = u'网址'
                    com_sheet['D1'] = u'电话'
                    com_sheet['E1'] = u'邮箱'
                    com_sheet['F1'] = u'地址'
                    com_sheet['G1'] = u'融资历史'
                    com_sheet.column_dimensions['B'].width = 50
                    com_sheet.column_dimensions['G'].width = 80
                    ws_com_hang = 2
                    for com in com_list:
                        com_events = MergeFinanceData.objects.filter(com_id=com.com_id)
                        com_event_list = []
                        for com_event in com_events:
                            if com_event.investormerge == 1:
                                invest_with_list = []
                                if hasattr(com_event.invsest_with, '__iter__'):
                                    for invesdic in com_event.invsest_with:
                                        if isinstance(invesdic, dict):
                                            invest_with_list.append(invesdic.get('invst_name'))
                                        if isinstance(invesdic, unicode):
                                            invest_with_list.append(invesdic)
                                invest_with_str = ','.join(invest_with_list)
                            else:
                                invest_with_str = com_event.merger_with
                            com_event_list.append('轮次：%s, 行业：%s->%s , 日期：%s , 投资方：%s, 投资金额：%s' % (
                                com_event.round, com_event.com_sub_cat_name, com_event.com_cat_name, com_event.date,
                                invest_with_str, com_event.money))
                        com_eventstr = '\n\r'.join(com_event_list)
                        com_sheet['A%s' % ws_com_hang] = str(com.com_name)
                        com_sheet['B%s' % ws_com_hang] = str(com.com_des) if com.com_des else '暂无'
                        com_sheet['C%s' % ws_com_hang] = str(com.com_web) if com.com_web else '暂无'
                        com_sheet['D%s' % ws_com_hang] = str(com.mobile) if com.mobile else '暂无'
                        com_sheet['E%s' % ws_com_hang] = str(com.email) if com.email else '暂无'
                        com_sheet['F%s' % ws_com_hang] = str(com.com_addr) if com.com_addr else '暂无'
                        com_sheet['G%s' % ws_com_hang] = com_eventstr if len(com_eventstr) > 0 else '暂无'
                        ws_com_hang += 1
                ws_org_hang += 1
            wb.save('/Users/investarget/Desktop/django_server/test.xlsx')

    d = downloadAllDataroomFile(org_qs, path)
    d.start()
